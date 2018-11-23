import datetime
from typing import List

from django.test import TestCase
from django.urls import reverse
from openpyxl import load_workbook

from datamap.models import Datamap
from excelparser.tests.factories.datamap_factories import DatamapFactory
from excelparser.tests.factories.datamap_factories import DatamapLineFactory
from excelparser.tests.factories.datamap_factories import ProjectFactory
from register.models import FinancialQuarter
from register.models import Project


class FactoryTests(TestCase):
    """
    These are tests to explore the Factory Boy API.
    """

    def setUp(self):
        self.datamap = DatamapFactory()
        self.project = ProjectFactory()
        self.datamapline = DatamapLineFactory(datamap=self.datamap)

    def test_datamap_factory(self):
        self.assertEqual(self.datamap.name, "Test Datamap from Factory")

    def test_datamapline_factory(self):
        self.assertEqual(self.datamapline.datamap, self.datamap)
        self.assertEqual(self.datamapline.key, "Test key")

    def test_project_factory(self):
        self.assertEqual(self.project.name, "Test Project")
        self.assertEqual(self.project.project_type.name, "Test ProjectType")
        self.assertEqual(self.project.stage.name, "Test Stage")


class ParsedSpreadsheet:
    def __init__(
        self,
        template_path: str,
        project: Project,
        fq: FinancialQuarter,
        datamap: Datamap,
    ):
        self.template_path = template_path
        self._project = project
        self.fq = fq
        self.datamap = datamap
        self.sheet_data = []

        self._get_sheets()

    def _process_sheets(self):
        wb = load_workbook(self.template_path)
        for ws in self.sheets:
            self.sheet_data.append(wb[ws])

    def process(self):
        self._process_sheets()

    @property
    def project_name(self) -> str:
        return self._project.name

    def _get_sheets(self) -> None:
        """
        Get the names of the sheets in an Excel file.
        :return:
        :rtype:
        """
        wb = load_workbook(self.template_path)
        self.sheets = wb.sheetnames


class ExcelParserIntegrationTests(TestCase):
    def setUp(self):
        self.financial_quarter = FinancialQuarter.objects.create(quarter=4, year=2018)
        self.project = ProjectFactory()
        self.datamap = DatamapFactory()
        DatamapLineFactory(key="Project Name", sheet="Test Sheet", cell_ref="B1")
        DatamapLineFactory(key="Total Cost", sheet="Test Sheet", cell_ref="B2")
        DatamapLineFactory(key="SRO", sheet="Test Sheet", cell_ref="B3")
        DatamapLineFactory(key="SRO Retirement Date", sheet="Test Sheet", cell_ref="B4")
        self.populated_template = "/home/lemon/code/python/dbasik-dev/dbasik-dftgovernance/excelparser/tests/populated.xlsm"

    def test_can_get_to_populated_template_upload_page(self):
        response = self.client.get(reverse("excelparser:process_populated"))
        self.assertEqual(response.status_code, 200)
        self.assertTemplateUsed(response, "excelparser/process_populated_template.html")

    def test_can_post_populated_template_form(self):
        response = self.client.post(
            reverse("excelparser:process_populated"),
            data={
                "datamap": self.datamap,
                "project": self.project,
                "financial_quarter": self.financial_quarter,
                "source_file": self.populated_template,
            },
        )
        self.assertEqual(response.status_code, 200)

    def test_parsed_spreadsheet_for_single_project(self):
        parsed_spreadsheet = ParsedSpreadsheet(
            template_path=self.populated_template,
            project=self.project,
            fq=self.financial_quarter,
            datamap=self.datamap,
        )
        self.assertEqual(parsed_spreadsheet.sheets, ["Test Sheet 1", "Test Sheet 2"])
        self.assertEqual(parsed_spreadsheet.project_name, "Test Project")
        # self.assertEqual(parsed_spreadsheet['Test Sheet 1'])
        # self.assertEqual(parsed_spreadsheet["Total Cost"], 45.2)
        # self.assertEqual(parsed_spreadsheet["SRO"], "John Milton")
        # self.assertEqual(
        #     parsed_spreadsheet["SRO Retirement Date"], datetime.date(2022, 2, 23)
        # )

    def test_individual_excel_sheet(self):
        parsed_spreadsheet = ParsedSpreadsheet(
            template_path=self.populated_template,
            project=self.project,
            fq=self.financial_quarter,
            datamap=self.datamap,
        )
        parsed_spreadsheet.process()
        test_sheet_1_data = parsed_spreadsheet.sheet_data[0]
        self.assertEqual(test_sheet_1_data.title, "Test Sheet 1")
        self.assertEqual(test_sheet_1_data['B1'].value, "Testable Project")
        self.assertEqual(test_sheet_1_data['B2'].value, 45.2)
        self.assertEqual(test_sheet_1_data['B3'].value, "John Milton")
        # do we want this to be a datetime object??
        self.assertEqual(test_sheet_1_data['B4'].value, datetime.datetime(2022, 2, 23, 0, 0))


