import datetime
from inspect import getmembers
from typing import List, Tuple, Union

from dbasik.datamap.models import Datamap
from dbasik.register.models import FinancialQuarter
from dbasik.returns.models import ReturnItem
from openpyxl import Workbook, utils

RETURN_ITEM_PARM_STRS = [
    "value_str",
    "value_float",
    "value_date",
    "value_datetime",
    "value_int",
]


def _get_populated_param(
    return_item: ReturnItem,
) -> Tuple[str, Union[str, int, float, datetime.datetime, datetime.date]]:
    _ri_params = [
        param for param in getmembers(return_item) if param[0] in RETURN_ITEM_PARM_STRS
    ]
    _non_none_params = [
        param for param in _ri_params if param[1] is not None and param[1] != ""
    ]
    if len(_non_none_params) > 1:
        raise ValueError(
            "You can't have multiple populated params in a ReturnItem object"
        )
    else:
        try:
            _non_none_params[0]
        except IndexError:
            # all params are None, therefore empty cell
            # we can just pass it as value_str
            return ("value_str", "")
        return _non_none_params[0]


def generate_master(
    financial_quarter: FinancialQuarter, output: str, datamap: Datamap
) -> None:
    # we need to get all returns for a FinancialQuarter
    all_returns = financial_quarter.return_financial_quarters.all().order_by("project")
    keys: List = [dml.key for dml in datamap.datamaplines.all().order_by("pk")]
    wb = Workbook()
    ws = wb.active

    # do the key column
    for row in range(1, len(keys) + 1):
        ws.cell(column=1, row=row, value=keys[row - 1])

    column_counter = 2  # start at column 2 after the key column
    for ret in all_returns:
        ret_items = ret.return_returnitems.all().order_by("datamapline")
        for counter_through_items, item in enumerate(ret_items, start=1):
            param: str = _get_populated_param(item)[0]
            ws.cell(
                column=column_counter,
                row=counter_through_items,
                value=item.__dict__[param],
            )
        column_counter += 1
    ws.title = "Master Data"
    dest_filename = output
    wb.save(filename=dest_filename)
