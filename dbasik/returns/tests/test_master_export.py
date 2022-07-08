import datetime
import os

from django.test import TestCase

from dbasik.datamap.models import DatamapLine
from factories.datamap_factories import DatamapFactory, ProjectFactory
from dbasik.register.models import FinancialQuarter
from returns.models import Return, ReturnItem
from returns.helpers import generate_master

from openpyxl import load_workbook
from openpyxl import utils


class TestMasterExport(TestCase):
    def setUp(self):
        self.output_file = "master.xlsm"
        self.project1 = ProjectFactory(name="Test Project 1")
        self.project2 = ProjectFactory(name="Test Project 2")
        self.fq = FinancialQuarter.objects.create(quarter=1, year=2010)

        self.return_obj1 = Return.objects.create(
            project=self.project1, financial_quarter=self.fq
        )
        self.return_obj2 = Return.objects.create(
            project=self.project2, financial_quarter=self.fq
        )

        self.datamap = DatamapFactory()
        self.dml1 = DatamapLine.objects.create(
            datamap=self.datamap,
            key="Test dml key1_str",
            sheet="Test Sheet 1",
            cell_ref="A1",
        )
        self.dml2 = DatamapLine.objects.create(
            datamap=self.datamap,
            key="Test dml key2_str",
            sheet="Test Sheet 1",
            cell_ref="A2",
        )
        self.dml3 = DatamapLine.objects.create(
            datamap=self.datamap,
            key="Test dml key3_date",
            sheet="Test Sheet 1",
            cell_ref="A3",
        )

        self.ri1 = ReturnItem.objects.create(
            parent=self.return_obj1,
            datamapline=self.dml1,
            value_str="Test dml return_ob1 str value1",
        )
        self.ri2 = ReturnItem.objects.create(
            parent=self.return_obj1,
            datamapline=self.dml2,
            value_str="Test dml return_ob1 str value2",
        )
        self.ri3 = ReturnItem.objects.create(
            parent=self.return_obj1,
            datamapline=self.dml3,
            value_date=datetime.date(2010, 10, 10),
        )
        self.ri4 = ReturnItem.objects.create(
            parent=self.return_obj2,
            datamapline=self.dml1,
            value_str="Test dml return_ob2 str value1",
        )
        self.ri5 = ReturnItem.objects.create(
            parent=self.return_obj2,
            datamapline=self.dml2,
            value_str="Test dml return_ob2 str value2",
        )
        self.ri6 = ReturnItem.objects.create(
            parent=self.return_obj2,
            datamapline=self.dml3,
            value_date=datetime.date(2011, 7, 12),
        )

    def test_setup(self):
        self.assertEqual(self.return_obj1.project.name, "Test Project 1")
        self.assertEqual(self.return_obj2.project.name, "Test Project 2")
        self.assertEqual(self.datamap.name, "Test Datamap from Factory")
        self.assertEqual(self.datamap.datamaplines.first().key, "Test dml key1_str")
        self.assertEqual(
            self.return_obj1.return_returnitems.first().datamapline, self.dml1
        )
        self.assertEqual(
            self.return_obj1.return_returnitems.first().value_str,
            "Test dml return_ob1 str value1",
        )
        self.assertEqual(
            self.return_obj1.return_returnitems.all()[2].value_date,
            datetime.date(2010, 10, 10),
        )

    def test_master_output(self):
        generate_master(
            financial_quarter=self.fq, output=self.output_file, datamap=self.datamap
        )
        wb = load_workbook("master.xlsm", data_only=True)
        ws = wb.active
        self.assertEqual(ws.title, "Master Data")
        self.assertEqual(ws["A1"].value, "Test dml key1_str")
        self.assertEqual(ws["B1"].value, "Test dml return_ob1 str value1")
        self.assertEqual(ws["C1"].value, "Test dml return_ob2 str value1")
        self.assertEqual(ws["A2"].value, "Test dml key2_str")
        self.assertEqual(ws["B2"].value, "Test dml return_ob1 str value2")
        self.assertEqual(ws["C2"].value, "Test dml return_ob2 str value2")
        self.assertEqual(ws["A3"].value, "Test dml key3_date")
        # dates always come out of Excel as datetime objects!
        self.assertEqual(ws["B3"].value, datetime.datetime(2010, 10, 10, 0, 0))
        self.assertEqual(ws["C3"].value, datetime.datetime(2011, 7, 12, 0, 0))
        os.remove(self.output_file)
